#!/usr/bin/env python3
"""Parse soumu_shikuchoson_202601.xlsx and update standardization.json.

This Excel has a different structure from the original shikucho_progress.xlsx:
  - Sheet1 '令和8年1月_進捗状況': Prefecture-level summary (47 rows)
    col0=都道府県, col1=完了率, col2+=業務別完了率
  - Sheet2 '業務別ステップ別進捗一覧': Step-level detail per prefecture x business
    col0=都道府県名, col1=市区町村名, col2=業務名, col3-42=40 steps

Strategy:
  - Load existing standardization.json (produced by parse_soumu.py from 1,741-municipality data)
  - Enrich it with step-level detail from Sheet2
  - Update prefecture-level rates from Sheet1 if available
  - Write back to standardization.json
"""
from __future__ import annotations

import json
import statistics
from pathlib import Path

EXCEL_FILE = Path(__file__).parent / "data/gov_downloads/soumu_shikuchoson_202601.xlsx"
OUT_DIR = Path(__file__).parent.parent / "public/data"
EXISTING_JSON = OUT_DIR / "standardization.json"

# Total steps per business in the step-detail sheet
TOTAL_STEPS = 40

# Statuses that count as "done" for completion rate calculation
DONE_STATUSES = {"完了済み", "対象外"}


def _pct(v):
    if v is None:
        return None
    try:
        return round(float(v), 4)
    except (ValueError, TypeError):
        return None


def _parse_sheet1(wb):
    """Parse prefecture-level summary from Sheet1."""
    ws = wb["令和8年1月_進捗状況"]
    rows = list(ws.iter_rows(values_only=True))

    # Row 1: header with business names merged
    # Row 2: sub-header with individual business names
    # Row 3+: data
    header_row = rows[0]
    sub_header = rows[1]

    # Collect business names from sub-header (col2+)
    businesses = []
    for i in range(2, len(sub_header)):
        name = sub_header[i]
        if name and str(name).strip():
            businesses.append(str(name).strip())
    # Also check header row for business names in merged cells
    if not businesses:
        for i in range(2, len(header_row)):
            name = header_row[i]
            if name and str(name).strip():
                businesses.append(str(name).strip())

    prefectures = []
    for row in rows[2:]:
        if not row[0]:
            continue
        pref_name = str(row[0]).strip()
        overall_rate = _pct(row[1])
        biz_rates = {}
        for j, biz in enumerate(businesses):
            col_idx = 2 + j
            if col_idx < len(row):
                biz_rates[biz] = _pct(row[col_idx])
        prefectures.append({
            "prefecture": pref_name,
            "overall_rate": overall_rate,
            "business_rates": biz_rates,
        })

    return prefectures, businesses


def _parse_sheet2(wb):
    """Parse step-level detail from Sheet2.

    Returns list of dicts:
      {prefecture, city, business, steps: [{name, status, done}], completion_rate}
    """
    sheet_name = "業務別ステップ別進捗一覧"
    if sheet_name not in wb.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' not found, skipping step detail")
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Row 1 is header with step names in col3+
    header = rows[0]
    step_names = []
    for i in range(3, len(header)):
        name = header[i]
        if name:
            step_names.append(str(name).strip())
        else:
            step_names.append(f"Step_{i-2}")

    details = []
    for row in rows[1:]:
        if not row[0]:
            continue
        pref = str(row[0]).strip()
        city = str(row[1]).strip() if row[1] else ""
        business = str(row[2]).strip() if row[2] else ""

        steps = []
        done_count = 0
        total_count = 0
        for j, sname in enumerate(step_names):
            col_idx = 3 + j
            status = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] else "未着手"
            is_done = status in DONE_STATUSES
            if is_done:
                done_count += 1
            total_count += 1
            steps.append({
                "name": sname,
                "status": status,
                "done": is_done,
            })

        completion_rate = round(done_count / total_count, 4) if total_count > 0 else 0.0
        details.append({
            "prefecture": pref,
            "city": city,
            "business": business,
            "steps": steps,
            "done_count": done_count,
            "total_steps": total_count,
            "completion_rate": completion_rate,
        })

    return details


def run():
    import openpyxl

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load existing standardization.json
    if EXISTING_JSON.exists():
        with open(EXISTING_JSON, "r", encoding="utf-8") as f:
            existing = json.load(f)
        print(f"Loaded existing standardization.json: {existing['summary']['total']} municipalities")
    else:
        print("Warning: No existing standardization.json found. Creating new one from prefecture data only.")
        existing = None

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    # Parse both sheets
    pref_summary, sheet1_businesses = _parse_sheet1(wb)
    step_details = _parse_sheet2(wb)
    wb.close()

    print(f"Sheet1: {len(pref_summary)} prefectures, businesses: {sheet1_businesses}")
    print(f"Sheet2: {len(step_details)} step-detail rows")

    # Build step detail summary by prefecture
    pref_step_map = {}
    for d in step_details:
        pref = d["prefecture"]
        if pref not in pref_step_map:
            pref_step_map[pref] = []
        pref_step_map[pref].append({
            "business": d["business"],
            "city": d["city"],
            "completion_rate": d["completion_rate"],
            "done_count": d["done_count"],
            "total_steps": d["total_steps"],
            "steps": d["steps"],
        })

    # Build step completion summary by step name (across all prefectures/businesses)
    step_agg = {}
    for d in step_details:
        for s in d["steps"]:
            name = s["name"]
            if name not in step_agg:
                step_agg[name] = {"total": 0, "done": 0}
            step_agg[name]["total"] += 1
            if s["done"]:
                step_agg[name]["done"] += 1

    step_completion = []
    for name, counts in step_agg.items():
        rate = round(counts["done"] / counts["total"], 4) if counts["total"] > 0 else 0.0
        step_completion.append({
            "step_name": name,
            "done_count": counts["done"],
            "total": counts["total"],
            "completion_rate": rate,
        })

    # Build output
    if existing:
        out = existing.copy()
    else:
        # Fallback: build from prefecture-level data only
        rates = [p["overall_rate"] for p in pref_summary if p["overall_rate"] is not None]
        out = {
            "summary": {
                "data_month": "2026-01",
                "deadline": "2026-03-31",
                "total": len(pref_summary),
                "avg_rate": round(statistics.mean(rates), 4) if rates else 0,
                "median_rate": round(statistics.median(rates), 4) if rates else 0,
                "completed_count": sum(1 for r in rates if r >= 1.0),
                "critical_count": sum(1 for r in rates if r < 0.5),
                "at_risk_count": sum(1 for r in rates if 0.5 <= r < 0.8),
                "on_track_count": sum(1 for r in rates if 0.8 <= r < 1.0),
            },
            "prefectures": [{
                "prefecture": p["prefecture"],
                "avg_rate": p["overall_rate"],
                "count": 1,
                "completed": 1 if p["overall_rate"] and p["overall_rate"] >= 1.0 else 0,
                "critical": 1 if p["overall_rate"] and p["overall_rate"] < 0.5 else 0,
            } for p in pref_summary],
            "businesses": [],
            "risk_municipalities": [],
            "municipalities": [],
        }

    # Enrich: add step_detail section
    out["step_detail"] = {
        "source": "総務省 市区町村の標準化・共通化に係る手順書の各ステップの進捗状況（令和8年1月時点）",
        "sheet_businesses": sheet1_businesses,
        "prefecture_steps": pref_step_map,
        "step_completion_summary": step_completion,
    }

    # Enrich: update prefecture rates from Sheet1 if we have a match
    if existing and pref_summary:
        pref_rate_map = {p["prefecture"]: p for p in pref_summary}
        for pref_entry in out.get("prefectures", []):
            match = pref_rate_map.get(pref_entry["prefecture"])
            if match and match.get("overall_rate") is not None:
                pref_entry["soumu_overall_rate"] = match["overall_rate"]
                pref_entry["soumu_business_rates"] = match.get("business_rates", {})

    # Write output
    with open(EXISTING_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    total_steps_done = sum(d["done_count"] for d in step_details)
    total_steps_all = sum(d["total_steps"] for d in step_details)
    overall_step_rate = total_steps_done / total_steps_all if total_steps_all > 0 else 0
    print(
        f"Done: {len(pref_summary)} prefectures, "
        f"{len(step_details)} step rows, "
        f"step completion: {overall_step_rate:.1%}"
    )
    print(f"Output: {EXISTING_JSON}")


if __name__ == "__main__":
    run()
