#!/usr/bin/env python3
"""
KWプランナーXLSX自動更新スクリプト
記事公開時にステータスを「公開済み」に更新する

使用法:
  python3 update_kw_planner.py --slug gc-cost-structural-factors --title "移行コストが3〜5倍に膨らむ5つの原因" --date 2026-03-20
  python3 update_kw_planner.py --slug gc-cost-structural-factors  # タイトル省略可（slug でメモ欄を検索）
"""

import argparse
import glob
import os
import sys
import shutil
import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from difflib import SequenceMatcher
except ImportError:
    print("ERROR: openpyxl が必要です。pip3 install openpyxl を実行してください。")
    sys.exit(1)

# 色定義
FILL_SEO  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 緑
FILL_NOTE = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # 青

def _default_planner_path() -> str:
    """最新のKWプランナーXLSXを更新日時順で取得する（バージョン番号ではなくタイムスタンプ優先）"""
    pattern = str(Path.home() / "Library/CloudStorage/GoogleDrive-tadashi.kudo@reraflow.com/マイドライブ/drive-workspace/contents/PJ19/gcportal_kw_planner_v*.xlsx")
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"KWプランナーXLSXが見つかりません: {pattern}")
    return files[0]

XLSX_PATH = Path(os.environ.get("KW_PLANNER_PATH", _default_planner_path()))

def similarity(a: str, b: str) -> float:
    """文字列類似度（0〜1）"""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, str(a).lower(), str(b).lower()).ratio()

def find_matching_row(ws, slug, title, headers):
    """slugまたはタイトルでマッチする行を探す"""
    status_col = headers.index('ステータス') + 1
    title_col = headers.index('推奨記事タイトル') + 1
    memo_col = headers.index('メモ') + 1
    kw_col = headers.index('キーワード') + 1

    best_row = None
    best_score = 0.0

    for r in range(2, ws.max_row + 1):
        current_status = ws.cell(r, status_col).value
        # すでに公開済みの行はスキップ（slugが一致する場合は更新）
        memo = str(ws.cell(r, memo_col).value or "")
        row_title = str(ws.cell(r, title_col).value or "")
        row_kw = str(ws.cell(r, kw_col).value or "")

        # slugがメモ欄に含まれていれば完全一致
        if slug and slug in memo:
            return r

        # タイトル類似度チェック
        if title:
            score = similarity(title, row_title)
            if score > best_score:
                best_score = score
                best_row = r

        # キーワードがslugに含まれているかチェック
        if slug and row_kw:
            kw_clean = row_kw.replace(" ", "-").replace("　", "-").lower()
            if kw_clean in slug or slug in kw_clean:
                score = 0.7
                if score > best_score:
                    best_score = score
                    best_row = r

    if best_score >= 0.5:
        return best_row

    return None

def update_kw_planner(slug, title=None, date=None, dry_run=False):
    if not XLSX_PATH.exists():
        print(f"ERROR: ファイルが見つかりません: {XLSX_PATH}")
        sys.exit(1)

    pub_date = date or datetime.date.today().isoformat()

    # バックアップ（_old_kw_planner/ に保存）
    if not dry_run:
        old_dir = XLSX_PATH.parent / '_old_kw_planner'
        old_dir.mkdir(exist_ok=True)
        backup_path = old_dir / (XLSX_PATH.stem + f"_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(XLSX_PATH, backup_path)
        print(f"バックアップ: {backup_path.name}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['KWリスト']

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    status_col = headers.index('ステータス') + 1
    date_col = headers.index('公開予定') + 1
    memo_col = headers.index('メモ') + 1
    title_col = headers.index('推奨記事タイトル') + 1

    row = find_matching_row(ws, slug, title, headers)

    if row is None:
        print(f"WARNING: マッチする行が見つかりませんでした (slug={slug}, title={title})")
        print("  手動でXLSXを更新してください。")
        sys.exit(1)

    row_title = ws.cell(row, title_col).value
    current_status = ws.cell(row, status_col).value
    print(f"マッチ: 行{row} 「{row_title}」 (現状: {current_status})")

    if dry_run:
        print(f"[DRY RUN] ステータス → 公開済み, 公開予定 → {pub_date}, メモにslug追記: {slug}")
        return

    # 更新
    ws.cell(row, status_col).value = '公開済み'
    ws.cell(row, date_col).value = pub_date

    # メモにslugを追記（既存メモを保持）
    current_memo = ws.cell(row, memo_col).value or ""
    if slug not in str(current_memo):
        new_memo = f"{current_memo} | slug:{slug}".strip(" |")
        ws.cell(row, memo_col).value = new_memo

    # 色付け（配信先に応じて行全体を塗る）
    dist_col = headers.index('配信先') + 1
    dist = ws.cell(row, dist_col).value
    fill = FILL_SEO if dist == 'SEO' else FILL_NOTE if dist == 'note' else None
    if fill:
        for c in range(1, ws.max_column + 1):
            ws.cell(row, c).fill = fill

    # 集計サマリーシートを再計算（公開済み件数）
    if '集計サマリー' in wb.sheetnames:
        ws_summary = wb['集計サマリー']
        published_count = sum(
            1 for r in range(2, ws.max_row + 1)
            if ws.cell(r, status_col).value == '公開済み'
        )
        for r in range(2, ws_summary.max_row + 1):
            label = ws_summary.cell(r, 1).value
            if label and '公開済み' in str(label):
                ws_summary.cell(r, 2).value = published_count
                break

    wb.save(XLSX_PATH)
    print(f"✅ 更新完了: 行{row} → 公開済み ({pub_date}) | slug:{slug}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='KWプランナーXLSX更新')
    parser.add_argument('--slug', required=True, help='記事のslug')
    parser.add_argument('--title', help='記事タイトル（マッチ精度向上）')
    parser.add_argument('--date', help='公開日 (YYYY-MM-DD, デフォルト: 今日)')
    parser.add_argument('--dry-run', action='store_true', help='変更せず確認のみ')
    args = parser.parse_args()

    update_kw_planner(
        slug=args.slug,
        title=args.title,
        date=args.date,
        dry_run=args.dry_run
    )
