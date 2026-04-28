#!/usr/bin/env python3
"""
高齢化率・財政比率データを管理APIで投入するためのUPDATE SQLを生成する
出力: /tmp/aging_update.sql
"""

import openpyxl, json, sys

AGING_PATH        = "/tmp/muni_aging/001023719.xlsx"
FISCAL_INDEX_PATH = "/tmp/muni_finance/fiscal_index.xlsx"

PREF_CODE_MAP = {
    "01":"北海道","02":"青森県","03":"岩手県","04":"宮城県","05":"秋田県",
    "06":"山形県","07":"福島県","08":"茨城県","09":"栃木県","10":"群馬県",
    "11":"埼玉県","12":"千葉県","13":"東京都","14":"神奈川県","15":"新潟県",
    "16":"富山県","17":"石川県","18":"福井県","19":"山梨県","20":"長野県",
    "21":"岐阜県","22":"静岡県","23":"愛知県","24":"三重県","25":"滋賀県",
    "26":"京都府","27":"大阪府","28":"兵庫県","29":"奈良県","30":"和歌山県",
    "31":"鳥取県","32":"島根県","33":"岡山県","34":"広島県","35":"山口県",
    "36":"徳島県","37":"香川県","38":"愛媛県","39":"高知県","40":"福岡県",
    "41":"佐賀県","42":"長崎県","43":"熊本県","44":"大分県","45":"宮崎県",
    "46":"鹿児島県","47":"沖縄県",
}

def load_aging():
    wb = openpyxl.load_workbook(AGING_PATH)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(values_only=True):
        code = row[0]
        if not code or not str(code).replace(" ", "").isdigit():
            continue
        if str(row[3]).strip() != "計":
            continue
        pref_code = str(code).strip().zfill(6)[:2]
        pref = PREF_CODE_MAP.get(pref_code, "")
        if not pref:
            continue
        city = str(row[2]).strip() if row[2] else ""
        if not city or city == "計":
            continue
        total = row[4]
        if not total or total == 0:
            continue
        aged = sum(v for v in row[18:26] if v and isinstance(v, (int, float)))
        result[(pref, city)] = round(aged / float(total) * 100, 1)
    print(f"[aging] {len(result)}件", file=sys.stderr)
    return result

def load_ratios():
    wb = openpyxl.load_workbook(FISCAL_INDEX_PATH)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(values_only=True):
        code = row[0]
        if not code or not str(code).isdigit():
            continue
        pref = str(row[1]).strip()
        city = str(row[2]).strip()
        def sf(v):
            if v is None: return None
            s = str(v).strip()
            return None if s in ("-", "", "－") else float(s)
        result[(pref, city)] = (sf(row[4]), sf(row[5]), sf(row[6]))
    print(f"[ratios] {len(result)}件", file=sys.stderr)
    return result

aging  = load_aging()
ratios = load_ratios()

# Merge by (pref, city)
keys = set(aging.keys()) | set(ratios.keys())

rows = []
for (pref, city) in sorted(keys):
    ar = aging.get((pref, city))
    r  = ratios.get((pref, city), (None, None, None))
    cer, rdr, fbr = r
    def q(v):
        return "NULL" if v is None else str(v)
    esc_pref = pref.replace("'", "''")
    esc_city = city.replace("'", "''")
    rows.append(f"('{esc_pref}', '{esc_city}', {q(ar)}, {q(cer)}, {q(rdr)}, {q(fbr)})")

sql = f"""UPDATE municipalities AS m SET
  aging_rate = v.ar::numeric,
  current_expenditure_ratio = v.cer::numeric,
  real_debt_ratio = v.rdr::numeric,
  future_burden_ratio = v.fbr::numeric
FROM (VALUES
{chr(10).join('  ' + r + (',' if i < len(rows)-1 else '') for i, r in enumerate(rows))}
) AS v(prefecture, city, ar, cer, rdr, fbr)
WHERE m.prefecture = v.prefecture AND m.city = v.city;
"""

out = "/tmp/aging_update.sql"
with open(out, "w") as f:
    f.write(sql)

print(f"[done] {len(rows)}行 → {out}", file=sys.stderr)
# Print byte size
import os
print(f"[size] {os.path.getsize(out):,} bytes", file=sys.stderr)
