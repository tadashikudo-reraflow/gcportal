#!/usr/bin/env python3
"""
GCInsight → X Articles 投稿ステータス管理
- x_paste_ready=false な記事一覧表示・内容確認
- ※ 実際の投稿は手動（generate_x_paste.py でHTML生成→コピペ）
- ※ x_paste_ready フラグは generate_x_paste.py が自動でセット

使い方:
  python3 x_article_publisher.py --list    # X HTML未生成記事一覧表示
  python3 x_article_publisher.py --dry-run # 次の記事の内容確認のみ
"""

import os
import sys
import json
import argparse
import requests
import openpyxl
from datetime import datetime, date
from markdownify import markdownify as md

# ── 設定 ──
SUPABASE_URL = "https://msbwmfggvtyexvhmlifn.supabase.co"
SUPABASE_SERVICE_KEY = os.environ.get(
    "SUPABASE_SERVICE_ROLE_KEY",
    "***REDACTED_SUPABASE_SERVICE_ROLE***"
)
SITE_URL = "https://gcinsight.jp"
PDF_URL = "https://gcinsight.jp/report?from=nav"
KW_PLANNER_PATH = os.path.expandvars(
    "$GDRIVE_WORKSPACE/contents/PJ19/gcportal_kw_planner_v7_20260320.xlsx"
)
README_PATH = os.path.expandvars(
    "$GDRIVE_WORKSPACE/../../../マイドライブ/ObsidianVault/01_Projects/PJ19_GCInsight/README.md"
)

HEADERS = {
    "apikey": SUPABASE_SERVICE_KEY,
    "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    "Content-Type": "application/json",
}

CTA_TEMPLATE = """---

📊 全国1,741自治体のガバメントクラウド移行状況を無料で確認
👉 {site_url}

📄 移行コスト実態レポート（無料PDF）
👉 {pdf_url}
"""


def get_next_article():
    """X HTML未生成記事を id 昇順で1本取得"""
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/articles",
        headers=HEADERS,
        params={
            "select": "id,slug,title,content,content_format,cover_image,date",
            "is_published": "eq.true",
            "x_paste_ready": "eq.false",
            "order": "id.asc",
            "limit": 1,
        },
    )
    resp.raise_for_status()
    data = resp.json()
    return data[0] if data else None


def list_unposted():
    """X HTML未生成記事一覧"""
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/articles",
        headers=HEADERS,
        params={
            "select": "id,slug,title,date",
            "is_published": "eq.true",
            "x_paste_ready": "eq.false",
            "order": "id.asc",
        },
    )
    resp.raise_for_status()
    return resp.json()


def list_all_published():
    """公開済み記事の総数取得"""
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/articles",
        headers=HEADERS,
        params={
            "select": "id",
            "is_published": "eq.true",
        },
    )
    resp.raise_for_status()
    return len(resp.json())


def html_to_markdown(html_content):
    """HTML → Markdown変換（X Articles用）"""
    return md(
        html_content,
        heading_style="ATX",
        bullets="-",
        strip=["script", "style"],
    ).strip()


def build_x_article(article):
    """X Articles用コンテンツを組み立て"""
    raw_content = article.get("content", "")
    content_format = article.get("content_format", "markdown")
    # HTML形式の場合はMarkdownに変換
    if content_format == "html":
        content = html_to_markdown(raw_content)
    else:
        content = raw_content
    cta = CTA_TEMPLATE.format(site_url=SITE_URL, pdf_url=PDF_URL)
    cover_url = None
    if article.get("cover_image"):
        cover_url = f"{SITE_URL}{article['cover_image']}"
    return {
        "title": article["title"],
        "body": content + cta,
        "cover_image_url": cover_url,
    }


def mark_paste_ready(article_id):
    """Supabase の x_paste_ready フラグを更新"""
    resp = requests.patch(
        f"{SUPABASE_URL}/rest/v1/articles",
        headers={**HEADERS, "Prefer": "return=minimal"},
        params={"id": f"eq.{article_id}"},
        json={"x_paste_ready": True},
    )
    resp.raise_for_status()


def update_kw_planner(slug, posted_date_str=None):
    """KWプランナーの X Article 列を '作成済み' に更新"""
    if not os.path.exists(KW_PLANNER_PATH):
        print(f"⚠️  KWプランナーが見つかりません: {KW_PLANNER_PATH}")
        return

    wb = openpyxl.load_workbook(KW_PLANNER_PATH)
    ws = wb["KWリスト"]

    headers = [cell.value for cell in ws[1]]
    try:
        x_col = headers.index("X Article") + 1
        memo_col = headers.index("メモ") + 1
    except ValueError:
        print("⚠️  X Article 列が見つかりません")
        return

    updated = False
    for row in ws.iter_rows(min_row=2):
        memo = str(row[memo_col - 1].value or "")
        if f"slug:{slug}" in memo or slug in memo:
            row[x_col - 1].value = "作成済み"
            updated = True
            print(f"  KWプランナー更新: 行{row[0].row} slug={slug} → 作成済み")
            break

    if not updated:
        print(f"  ⚠️  KWプランナーに slug={slug} の行が見つかりませんでした（スキップ）")

    wb.save(KW_PLANNER_PATH)


def update_readme(total_count):
    """README の記事数を最新に更新"""
    if not os.path.exists(README_PATH):
        # 絶対パスで再試行
        alt = os.path.join(
            os.path.expanduser("~"),
            "Library/CloudStorage/GoogleDrive-tadashi.kudo@reraflow.com",
            "マイドライブ/ObsidianVault/01_Projects/PJ19_GCInsight/README.md",
        )
        if os.path.exists(alt):
            readme_path = alt
        else:
            print(f"⚠️  README が見つかりません")
            return
    else:
        readme_path = README_PATH

    with open(readme_path, "r", encoding="utf-8") as f:
        content = f.read()

    today = date.today().strftime("%Y-%m-%d")
    import re
    new_content = re.sub(
        r"- \*\*記事数\*\*: \d+本（.*?）",
        f"- **記事数**: {total_count}本（{today}時点）",
        content,
    )
    if new_content != content:
        with open(readme_path, "w", encoding="utf-8") as f:
            f.write(new_content)
        print(f"  README更新: {total_count}本 ({today})")


def main():
    parser = argparse.ArgumentParser(description="GCInsight → X Articles 投稿パイプライン")
    parser.add_argument("--dry-run", action="store_true", help="投稿せずに内容確認のみ")
    parser.add_argument("--list", action="store_true", help="未投稿記事一覧表示")
    args = parser.parse_args()

    if args.list:
        articles = list_unposted()
        print(f"X HTML未生成記事: {len(articles)}本\n")
        for a in articles:
            print(f"  [{a['id']}] {a['date']} {a['title']}")
        if not articles:
            print("✅ 全件のX HTML生成済みです。")
        return

    # 次の未処理記事を取得
    article = get_next_article()
    if not article:
        print("✅ X HTML未生成記事はありません。全件生成済みです。")
        return

    # X Articles用コンテンツ組み立て（確認用）
    payload = build_x_article(article)

    print(f"📄 次のX HTML生成対象:")
    print(f"  ID   : {article['id']}")
    print(f"  タイトル: {article['title']}")
    print(f"  Slug : {article['slug']}")
    print(f"  画像  : {payload['cover_image_url']}")
    print(f"\n--- 本文プレビュー（末尾500字）---")
    print(payload["body"][-500:])
    print("---\n")
    print("💡 generate_x_paste.py を実行してHTMLを生成してください")


if __name__ == "__main__":
    main()
