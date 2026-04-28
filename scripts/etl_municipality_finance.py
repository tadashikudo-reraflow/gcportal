#!/usr/bin/env python3
"""
自治体財政データETL
総務省ExcelからSupabase municipalitiesテーブルに財政プロフィールを投入する

データソース:
  fiscal_index.xlsx  : 全市町村の主要財政指標（令和5年度）- 財政力指数
  kessan.xlsx        : 市町村別決算状況調（令和6年度）  - 標準財政規模・人口

使い方:
  python3 scripts/etl_municipality_finance.py [--dry-run]
"""

import sys
import os
import openpyxl
import urllib.request
import json

DRY_RUN = "--dry-run" in sys.argv

SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "https://msbwmfggvtyexvhmlifn.supabase.co")
SERVICE_KEY  = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

FISCAL_INDEX_PATH = "/tmp/muni_finance/fiscal_index.xlsx"
KESSAN_PATH       = "/tmp/muni_finance/kessan.xlsx"

# 団体コード先頭2桁 → 都道府県名
PREF_CODE_MAP = {
    "01":"北海道","02":"青森県","03":"岩手県","04":"宮城県","05":"秋田県",
    "06":"山形県","07":"福島県","08":"茨城県","09":"栃木県","10":"群馬県",
    "11":"埼玉県","12":"千葉県","13":"東京都","14":"神奈川県","15":"新潟県",
    "16":"富山県","17":"石川県","18":"福井県","19":"山梨県","20":"長野県",
    "21":"岐阜県","22":"静岡県","23":"愛知県","24":"三重県","25":"滋賀県",
    "26":"京都府","27":"大阪府","28":"兵庫県","29":"奈良県","30":"和歌山県",
    "31":"鳥取県","32":"島根県","33":"岡山県","34":"広島県","35":"山口県",
    "36":"徳島県","37":"香川県","38":"愛媛県","39":"高知県","40":"福岡県",
    "41":"佐賀県","42":"長崎県","43":"熊本県","44":"大分県","45":"宮崎県",
    "46":"鹿児島県","47":"沖縄県",
}


def load_fiscal_index():
    """財政力指数ロード (fiscal_index.xlsx)
    Returns: dict[(pref, city)] = {fiscal_strength: float}
    """
    wb = openpyxl.load_workbook(FISCAL_INDEX_PATH)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(values_only=True):
        code = row[0]
        if not code or not str(code).isdigit():
            continue
        pref = str(row[1]).strip()
        city = str(row[2]).strip()
        fiscal_strength = row[3]
        result[(pref, city)] = {"fiscal_strength": float(fiscal_strength) if fiscal_strength else None}
    print(f"[fiscal_index] {len(result)}件ロード")
    return result


def load_kessan():
    """標準財政規模・人口ロード (kessan.xlsx)
    Returns: dict[(pref, city)] = {standard_fiscal_scale: int, population: int}
    """
    wb = openpyxl.load_workbook(KESSAN_PATH)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(values_only=True):
        code = row[14]
        if not code or not str(code).isdigit() or len(str(code)) != 6:
            continue
        pref_code = str(code)[:2]
        pref = PREF_CODE_MAP.get(pref_code)
        if not pref:
            continue
        city = str(row[15]).strip()
        population = row[16]
        std_fiscal = row[26]
        result[(pref, city)] = {
            "standard_fiscal_scale": int(std_fiscal) if std_fiscal and str(std_fiscal) != "-" else None,
            "population": int(population) if population and str(population) != "-" else None,
        }
    print(f"[kessan] {len(result)}件ロード")
    return result


def fetch_municipalities():
    """Supabaseから全municipalitiesを取得"""
    all_rows = []
    offset = 0
    limit = 1000
    while True:
        url = f"{SUPABASE_URL}/rest/v1/municipalities?select=id,prefecture,city&limit={limit}&offset={offset}"
        req = urllib.request.Request(url, headers={
            "apikey": SERVICE_KEY,
            "Authorization": f"Bearer {SERVICE_KEY}",
        })
        with urllib.request.urlopen(req) as r:
            batch = json.loads(r.read())
        all_rows.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
    print(f"[supabase] municipalities {len(all_rows)}件取得")
    return all_rows


def update_municipality(muni_id, payload):
    """1件UPDATE"""
    url = f"{SUPABASE_URL}/rest/v1/municipalities?id=eq.{muni_id}"
    body = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=body, method="PATCH", headers={
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    })
    with urllib.request.urlopen(req) as r:
        return r.status


def main():
    if not SERVICE_KEY:
        print("ERROR: SUPABASE_SERVICE_ROLE_KEY not set")
        sys.exit(1)

    fiscal = load_fiscal_index()
    kessan = load_kessan()
    munis  = fetch_municipalities()

    updated = 0
    no_match_fiscal = []
    no_match_kessan = []

    for m in munis:
        pref = m["prefecture"]
        city = m["city"]
        key  = (pref, city)

        payload = {"financial_data_year": "R5/R6"}

        f = fiscal.get(key)
        if f:
            payload["fiscal_strength"] = f["fiscal_strength"]
        else:
            no_match_fiscal.append(f"{pref} {city}")

        k = kessan.get(key)
        if k:
            payload["standard_fiscal_scale"] = k["standard_fiscal_scale"]
            payload["population"] = k["population"]
        else:
            no_match_kessan.append(f"{pref} {city}")

        if len(payload) <= 1:
            continue

        if DRY_RUN:
            print(f"  [DRY] {pref} {city}: {payload}")
        else:
            update_municipality(m["id"], payload)
        updated += 1

    print(f"\n--- 結果 ---")
    print(f"更新: {updated}件")
    print(f"fiscal_index未マッチ: {len(no_match_fiscal)}件")
    print(f"kessan未マッチ: {len(no_match_kessan)}件（町村は正常）")
    if no_match_fiscal:
        print("未マッチ（財政力指数）:", no_match_fiscal[:10])


if __name__ == "__main__":
    main()
