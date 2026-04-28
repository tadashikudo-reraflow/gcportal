#!/usr/bin/env python3
"""
自治体 高齢化率・財政比率 ETL
総務省ExcelからSupabase municipalitiesテーブルに投入する

データソース:
  001023719.xlsx  : 令和7年1月1日 住民基本台帳 年齢別人口（市区町村別）
  fiscal_index.xlsx : 全市町村の主要財政指標（令和5年度）

カラム:
  aging_rate               : 65歳以上人口 / 総人口 * 100
  current_expenditure_ratio: 経常収支比率 (fiscal_index col4)
  real_debt_ratio          : 実質公債費比率 (fiscal_index col5)
  future_burden_ratio      : 将来負担比率 (fiscal_index col6)

使い方:
  python3 scripts/etl_municipality_aging.py [--dry-run]
"""

import sys
import os
import openpyxl
import urllib.request
import json

DRY_RUN = "--dry-run" in sys.argv

SUPABASE_URL    = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "https://msbwmfggvtyexvhmlifn.supabase.co")
SERVICE_KEY     = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

AGING_PATH        = "/tmp/muni_aging/001023719.xlsx"
FISCAL_INDEX_PATH = "/tmp/muni_finance/fiscal_index.xlsx"

PREF_CODE_MAP = {
    "01":"北海道","02":"青森県","03":"岩手県","04":"宮城県","05":"秋田県",
    "06":"山形県","07":"福島県","08":"茨城県","09":"栃木県","10":"群馬県",
    "11":"埼玉県","12":"千葉県","13":"東京都","14":"神奈川県","15":"新潟県",
    "16":"富山県","17":"石川県","18":"福井県","19":"山梨県","20":"長野県",
    "21":"岐阜県","22":"静岡県","23":"愛知県","24":"三重県","25":"滋賀県",
    "26":"京都府","27":"大阪府","28":"兵庫県","29":"奈良県","30":"和歌山県",
    "31":"鳥取県","32":"島根県","33":"岡山県","34":"広島県","35":"山口県",
    "36":"徳島県","37":"香川県","38":"愛媛県","39":"高知県","40":"福岡県",
    "41":"佐賀県","42":"長崎県","43":"熊本県","44":"大分県","45":"宮崎県",
    "46":"鹿児島県","47":"沖縄県",
}


def load_aging_rate():
    """住民基本台帳から高齢化率を計算
    001023719.xlsx の構造:
      col0 = 団体コード, col1 = 都道府県名, col2 = 市区町村名
      col3 = 性別（計/男/女）, col4 = 総数
      col5-17 = 0~64歳（5歳区切り）
      col18-25 = 65-69, 70-74, 75-79, 80-84, 85-89, 90-94, 95-99, 100歳以上
    → col3 == "計" の行のみ使用
    """
    wb = openpyxl.load_workbook(AGING_PATH)
    ws = wb.active
    result = {}
    skipped = 0
    for row in ws.iter_rows(values_only=True):
        code = row[0]
        if not code or not str(code).replace(" ", "").isdigit():
            continue
        # 「計」行のみ（男/女は除外）
        gender = str(row[3]).strip() if row[3] else ""
        if gender != "計":
            continue

        pref_code = str(code).strip().zfill(6)[:2]
        pref = PREF_CODE_MAP.get(pref_code, "")
        if not pref:
            skipped += 1
            continue

        city = str(row[2]).strip() if row[2] else ""
        if not city or city in ("計", ""):
            continue

        total = row[4]
        if not total or total == 0:
            continue

        # 65歳以上: cols18-25 (0-indexed)
        aged = sum(v for v in row[18:26] if v and isinstance(v, (int, float)))
        rate = round(aged / float(total) * 100, 1)
        result[(pref, city)] = rate

    print(f"[aging] {len(result)}件ロード (skipped={skipped})")
    return result


def load_fiscal_ratios():
    """経常収支比率・実質公債費比率・将来負担比率 (fiscal_index.xlsx cols 4,5,6)"""
    wb = openpyxl.load_workbook(FISCAL_INDEX_PATH)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(values_only=True):
        code = row[0]
        if not code or not str(code).isdigit():
            continue
        pref = str(row[1]).strip()
        city = str(row[2]).strip()

        def safe_float(v):
            if v is None:
                return None
            s = str(v).strip()
            if s in ("-", "", "－"):
                return None
            try:
                return float(s)
            except Exception:
                return None

        result[(pref, city)] = {
            "current_expenditure_ratio": safe_float(row[4]),
            "real_debt_ratio":           safe_float(row[5]),
            "future_burden_ratio":       safe_float(row[6]),
        }
    print(f"[fiscal_ratios] {len(result)}件ロード")
    return result


def fetch_municipalities():
    all_rows = []
    offset = 0
    limit = 1000
    while True:
        url = f"{SUPABASE_URL}/rest/v1/municipalities?select=id,prefecture,city&limit={limit}&offset={offset}"
        req = urllib.request.Request(url, headers={
            "apikey": SERVICE_KEY,
            "Authorization": f"Bearer {SERVICE_KEY}",
        })
        with urllib.request.urlopen(req) as r:
            batch = json.loads(r.read())
        all_rows.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
    print(f"[supabase] municipalities {len(all_rows)}件取得")
    return all_rows


def update_municipality(muni_id, payload):
    url = f"{SUPABASE_URL}/rest/v1/municipalities?id=eq.{muni_id}"
    body = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=body, method="PATCH", headers={
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    })
    with urllib.request.urlopen(req) as r:
        return r.status


def main():
    if not SERVICE_KEY:
        print("ERROR: SUPABASE_SERVICE_ROLE_KEY not set")
        sys.exit(1)

    aging  = load_aging_rate()
    ratios = load_fiscal_ratios()
    munis  = fetch_municipalities()

    updated = 0
    no_aging  = []
    no_ratios = []

    for m in munis:
        pref = m["prefecture"]
        city = m["city"]
        key  = (pref, city)
        payload = {}

        a = aging.get(key)
        if a is not None:
            payload["aging_rate"] = a
        else:
            no_aging.append(f"{pref} {city}")

        r = ratios.get(key)
        if r:
            for col, val in r.items():
                if val is not None:
                    payload[col] = val
        else:
            no_ratios.append(f"{pref} {city}")

        if not payload:
            continue

        if DRY_RUN:
            print(f"  [DRY] {pref} {city}: {payload}")
        else:
            update_municipality(m["id"], payload)
        updated += 1

    print(f"\n--- 結果 ---")
    print(f"更新: {updated}件")
    print(f"aging未マッチ: {len(no_aging)}件")
    if no_aging[:10]:
        print("  未マッチ例:", no_aging[:10])
    print(f"fiscal_ratios未マッチ: {len(no_ratios)}件（町村は正常）")


if __name__ == "__main__":
    main()
